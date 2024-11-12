import openpyxl
wb = openpyxl.load_workbook("C://Users//lenovo//OneDrive//Documents//TestDataSubmit.xlsx")
ws = wb["Sheet1"]
print(ws.max_row)
cell = ws.cell(1,1)
print(cell.value)

def fetch_number_of_raws(sheetname):
    ws = wb[sheetname]
    return ws.max_row

def fetch_cell_data(sheetname, row, cell):
    ws = wb[sheetname]
    cell = ws.cell(int(row), int(cell))
    return cell.value
#print(fetch_number_of_raws("Sheet1"))
#print(fetch_cell_data("Sheet1", 1, 1))